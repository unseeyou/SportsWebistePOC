import plotly.graph_objs as go
import plotly.io as pio

import openpyxl as op
import random
from datetime import datetime, timedelta


def convert_to_24hr(time_str):
    # seperate the numbers from the letters to achieve proper format
    index = time_str.find(":") + 3  # account for minutes
    time_str = time_str[:index] + " " + time_str[index:]
    return datetime.strptime(time_str, "%I:%M %p")


def calculate_session_length(start_time, end_time):
    start_time = convert_to_24hr(start_time)
    end_time = convert_to_24hr(end_time)
    session_length = end_time - start_time
    # account for going over to the next day (edge case due to bad data)
    if session_length.total_seconds() < 0:
        session_length = (end_time + timedelta(days=1)) - start_time
    return session_length


def percent_bar_chart(data):
    sports = []
    attendance_percentages = []
    sorted_data = sorted(data, key=lambda x: x["Attendance"], reverse=True)

    for item in sorted_data:
        sports.append(item["Sport"])
        attendance_percentages.append(item["Attendance"])

    min_attendance = min(attendance_percentages)
    y_axis_start = max(0, min_attendance - 10)

    bar_chart = go.Figure(
        data=[go.Bar(x=sports, y=attendance_percentages, name="Attendance Percentage")]
    )

    bar_chart.update_layout(
        title="Sports Attendance",
        title_x=0.5,
        title_font_size=30,
        title_font_family="Arial",
        xaxis_title="Sport",
        yaxis_title="Attendance (%)",
        yaxis=dict(range=[y_axis_start, 100]),
        # clickmode="event+select"
    )

    return pio.to_json(bar_chart)


def demo_scatter_plot():
    x = [1, 9, 2, 4]
    y = [1, 2, 3, 4]
    fig = go.Figure(data=go.Scatter(x=x, y=y, mode="markers", name="Demo scatter plot"))
    scatter_plot = fig.to_html(full_html=False)
    return scatter_plot


def student_count_per_sport(fp: str):
    wb = op.load_workbook(fp)
    sheet = wb.active
    sports = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):  # skip the header row
        student_id = row[1]
        sport = row[12]

        if sport and student_id:
            if sport not in sports:
                sports[sport] = set()
            sports[sport].add(student_id)

    summary_data = []

    for sport, student_ids in sports.items():
        summary_data.append(
            {
                "sport": sport,
                "unique_students": len(student_ids),
            }
        )
    return summary_data


def average_session_length(fp: str):
    wb = op.load_workbook(fp)
    sheet = wb.active
    sessions = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):  # skip the header row
        start = row[15]
        end = row[16]
        sport = row[12]

        if start and end:
            session_length = calculate_session_length(start, end)
            if sport not in sessions:
                sessions[sport] = []
            sessions[sport].append(session_length.total_seconds() / 60)

    for sport, session_lengths in sessions.items():
        sessions[sport] = sum(session_lengths) / len(session_lengths)  # change total to average

    min_session_length = min(sessions.values())
    y_axis_start = max(0, min_session_length - 10)
    y_axis_end = max(sessions.values()) + 10

    chart = go.Figure(
        data=go.Bar(x=list(sessions.keys()), y=list(sessions.values()), name="Average Session Length"),
    )

    chart.update_layout(
        title="Average Session Length per Sport",
        title_x=0.5,
        title_font_size=30,
        title_font_family="Comfortaa",
        xaxis_title="Sport",
        yaxis_title="Average Session Length (min)",
        yaxis=dict(range=[y_axis_start, y_axis_end]),
        # clickmode="event+select"
    )

    return chart.to_html(full_html=False)


def cancelled_sessions(fp: str):
    wb = op.load_workbook(fp)
    sheet = wb.active
    sessions = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):  # skip the header row
        sport = row[12]
        cancelled_status = row[21]

        if cancelled_status:
            if sport not in sessions:
                sessions[sport] = {
                    "Total": 0,
                    "Cancelled": 0,
                    "Percentage": 0.0,
                }
            sessions[sport]["Total"] += 1
            if cancelled_status == "Yes":
                sessions[sport]["Cancelled"] += 1

    for sport in sessions:
        sessions[sport]["Percentage"] = round((sessions[sport]["Cancelled"] * 100) / sessions[sport]["Total"], 2)

    return sessions


def list_all_sports(fp: str):
    wb = op.load_workbook(fp)
    sheet = wb.active
    sports = []

    for row in sheet.iter_rows(min_row=2, values_only=True):  # skip the header row
        sport = row[12]

        if sport:
            if sport not in sports:
                sports.append(sport)

    return sports
